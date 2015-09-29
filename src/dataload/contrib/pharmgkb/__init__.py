# from .pharmgkb_parser import 
import json
import os
import requests
import urllib
from openpyxl import Workbook #2.2.3
from openpyxl import load_workbook
import re
from HTMLParser import HTMLParser
import itertools
from collections import OrderedDict

__METADATA__ = {
    "src_name": 'PharmGKB',
    "src_url": 'https://www.pharmgkb.org',
    "version": '0.1',
    "field": "pharmgkb"
}

infiles = ['https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1445556789&dlCls=HaplotypeSet&dlId=PA166128323&dlName=CPIC%20CYP2C19%20Haplotype%20Set',
'https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1446767521&dlCls=HaplotypeSet&dlId=PA165980499&dlName=CYP2D6%20Cytochrome%20P450%20Nomenclature%20DB%20Haplotype%20Set',
'https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1183805365&dlCls=HaplotypeSet&dlId=PA165980513&dlName=Haplotype%20Set%20PA165980513%20for%20DPYD',
'https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1445164739&dlCls=HaplotypeSet&dlId=PA166128346&dlName=CPIC%20TMPT%20Haplotype%20Set',
'https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1444880096&dlCls=HaplotypeSet&dlId=PA166115840&dlName=Haplotypes%20for%20UGT1A1%20(UGT%20Alleles%20Nomenclature%20page)'
]

def load_data():
    pass

''' To remove HTML tags from the Recommendation texts
'''
class MLStripper(HTMLParser):
    def __init__(self):
        self.reset()
        self.fed = []
    def handle_data(self, d):
        self.fed.append(d)
    def get_data(self):
        return ''.join(self.fed)
def strip_tags(html):
    s = MLStripper()
    s.feed(html)
    return s.get_data()

def getAllJson():
    fileOut = open('/Users/admin/Dropbox/Privat/00_Masterthesis/asdf.txt','w')

    rsidList = fillRsidList()
    for rs in rsidList:
        # getDosingGuidelineFromRsid(rs)
       fileOut.write(json.dumps(getDosingGuidelineFromRsid(rs), indent=4, sort_keys=True))
    # print json.dumps(getDosingGuidelineFromRsid('rs1801265'), indent=4, sort_keys=True)
    fileOut.close()

def fillRsidList():
    rsidList=[]
    for translationTablePerGene in os.listdir('/Users/admin/Dropbox/Privat/00_Masterthesis/pharmGkb_resources/'):
        if translationTablePerGene.endswith('.xlsx') and not translationTablePerGene.startswith('~'): #and translationTablePerGene.startswith(geneSymbolName):
            translationTablePerGeneWorkbook = load_workbook('/Users/admin/Dropbox/Privat/00_Masterthesis/pharmGkb_resources/'+translationTablePerGene,read_only=True)
            worksheetTranslationTablePerGene = translationTablePerGeneWorkbook.active
            for row in worksheetTranslationTablePerGene.rows:
                for cell in row:
                    if isinstance(cell.value,unicode): #dont parse datetime objects
                        if str(cell.value.encode('utf8','ignore')).strip().startswith('rs'): #we are searching for all rs id #encoding due to unicode characters, str(unicode) gives unicodeencdoerror
                            rsidList.append(str(cell.value.encode('utf8','ignore')).strip())
    print 'number of rs#s found: ',len(rsidList)
    return rsidList

def getGeneSymbolName(rsid):
    ''' Converts RsID to geneSymbolName through myvariant.info in iterations from 3 different annotation resources
    '''
    geneNameFromMyVariant = ''
    myvariantRsidRequest = requests.get('http://myvariant.info/v1/query?q='+rsid)
    if myvariantRsidRequest.status_code == requests.codes.ok:
        if bool(re.search('[r][s]\d+',rsid)):
            print 'http://myvariant.info/v1/query?q='+rsid
            commit_data = myvariantRsidRequest.json()
            try:
                print 'This is the hgvs id from myvariant.info:' ,commit_data['hits'][0]['_id']
            except IndexError:
                print 'invalid rs id ie no data found on myvariant.info'
                return
            print 'Searching for Genename...'
            try:
                geneNameFromMyVariant=commit_data['hits'][0]['dbsnp']['gene']['symbol']
            except KeyError:
                pass
                # print """KeyError: Key not found in ['hits'][0]['dbsnp']['gene']['symbol']"""
            except TypeError, e:
                print """TypeError: Multiple genes found in ['hits'][0]['dbsnp']['gene']['symbol']""",e
            else:
                print 'Genename found! This is the gene from myvariant.info:' ,commit_data['hits'][0]['dbsnp']['gene']['symbol']
            if geneNameFromMyVariant!='':
                return geneNameFromMyVariant
            try:
                geneNameFromMyVariant=commit_data['hits'][0]['snpeff']['ann'][0]['gene_name']
            except KeyError:
                pass
                # print """KeyError: Key not found in ['hits'][0]['snpeff']['ann'][0]['gene_name']"""
            else:
                print 'Genename found! This is the gene from myvariant.info:' ,commit_data['hits'][0]['snpeff']['ann'][0]['gene_name']
            if geneNameFromMyVariant!='':
                return geneNameFromMyVariant
            try:
                geneNameFromMyVariant=commit_data['hits'][0]['dbnsfp']['genename']
            except KeyError:
                pass
                # print """KeyError: Key not found in ['hits'][0]['dbnsfp']['genename']"""
            else:
                print 'Genename found! This is the gene from myvariant.info:' ,commit_data['hits'][0]['dbnsfp']['genename']
            if geneNameFromMyVariant!='':
                return geneNameFromMyVariant
            try:
                geneNameFromMyVariant=commit_data['hits'][0]['wellderly']['gene']
            except KeyError:
                pass
                # print """KeyError: Key not found in ['hits'][0]['dbnsfp']['genename']"""
            else:
                print 'Genename found! This is the gene from myvariant.info:' ,commit_data['hits'][0]['wellderly']['gene']
            if geneNameFromMyVariant!='' and not None and not type(geneNameFromMyVariant) is list:
                return geneNameFromMyVariant
            else:
                print 'genename not found on myvariant.info'
                raise ValueError('genename not found on myvariant.info or other error in genename search')
        else:
            print 'rsid malformed: '+rsid
            raise ValueError('rsid malformed: '+rsid)

    else:
        print '\nstatus_code at myvariant.info not ok!:',myvariantRsidRequest.headers['content-type']

def getHaplotypesFromTranslationtable(rsid):
    for translationTablePerGene in os.listdir('/Users/admin/Dropbox/Privat/00_Masterthesis/pharmGkb_resources/'):
        if translationTablePerGene.endswith('.xlsx') and not translationTablePerGene.startswith('~'):
            haplottypeListTemp=[]
            translationTablePerGeneWorkbook = load_workbook('/Users/admin/Dropbox/Privat/00_Masterthesis/pharmGkb_resources/'+translationTablePerGene,read_only=True)
            worksheetTranslationTablePerGene = translationTablePerGeneWorkbook.active
            coordinatesOfRsid = ''
            for row in worksheetTranslationTablePerGene.rows:
                for cell in row:
                    if isinstance(cell.value,unicode): #dont parse datetime objects
                        if str(cell.value.encode('utf8','ignore')).strip()==rsid: #encoding due to unicode characters, str(unicode) gives unicodeencdoerror
                            coordinatesOfRsid = cell.coordinate
            letterOfRsIdCell = ''
            if coordinatesOfRsid!='':
                letterOfRsIdCell = re.search('[A-Z]{1,2}', coordinatesOfRsid).group() #gives the letter of the coordinate
                rowCount = worksheetTranslationTablePerGene.get_highest_row()
                if not letterOfRsIdCell=='':
                    for i in range (1,rowCount+1):
                        try:
                            if worksheetTranslationTablePerGene[letterOfRsIdCell+str(i)].value: #take only non-empty cells
                                 if bool(re.search('\*\d',str(worksheetTranslationTablePerGene['B'+str(i)].value))): # pattern is star plus a digit then stop because we only want the basic star allels. We search in the B column because it contains the star alleles
                                    haplottypeListTemp.append(worksheetTranslationTablePerGene['B'+str(i)].value)
                        except IndexError, e:
                            print e
                        except:
                            pass
                print 'star alleles list:',haplottypeListTemp
                return haplottypeListTemp

def getStaralleleCombinations(haplottypeListComplete):
    starAllelesListTwoBasicTemp=[]
    if haplottypeListComplete is not None:
        for starAllele in itertools.combinations(haplottypeListComplete,2):
            starAllelesListTwoBasicTemp.append(starAllele)
        return starAllelesListTwoBasicTemp
    else:
        print 'no haplotypes found'

def getDosingGuidelineFromRsid(rsid):
    try:
        geneSymbolName = getGeneSymbolName(rsid)
    except ValueError:
        return
    if geneSymbolName == None:
        return
    haplottypeListComplete = getHaplotypesFromTranslationtable(rsid)
    starAllelesListTwoBasicCombinationsT = getStaralleleCombinations(haplottypeListComplete)
    print 'Searching for Dosing Guidelines for all ',len(starAllelesListTwoBasicCombinationsT), 'star allele combinations.'
    jsonSnp = OrderedDict()
    jsonSnp = {
        'rsid': rsid,
        'pharmgkb':
            {
                'drugrecommendations':[],
                'haplotypes':[]
            }
    }
    for starAllelesListTwoBasic in starAllelesListTwoBasicCombinationsT:
        if len(starAllelesListTwoBasic)>=2:
            # print 'Searching for Dosing Guidelines for star alleles:',starAllelesListTwoBasic[0],starAllelesListTwoBasic[1]
            for dosingGuidelinesJsonFile in os.listdir('/Users/admin/Dropbox/Privat/00_Masterthesis/pharmGkb_resources/dosingGuidelines.json/'):
                if dosingGuidelinesJsonFile.endswith('.json'):
                    if 'CPIC' in dosingGuidelinesJsonFile:
                        if geneSymbolName in dosingGuidelinesJsonFile:                        
                            '''  takes a json file and searches for the two given star alleles and tries to print all found dosing guidelines
                            '''
                            with open('/Users/admin/Dropbox/Privat/00_Masterthesis/pharmGkb_resources/dosingGuidelines.json/'+dosingGuidelinesJsonFile) as data_file:
                                parsedJsonFile = json.loads(data_file.read())
                            if 'guides' in parsedJsonFile:
                                for annsLoop in parsedJsonFile['guides'][0]['anns']:
                                    if 'location' in annsLoop:
                                        levelOfEvidence=''
                                        rec=''
                                        drug = ''
                                        for diplotypeLoop in annsLoop['location']['diplotypes']:
                                            # print diplotypeLoop
                                            if (diplotypeLoop['allele1']==starAllelesListTwoBasic[0] and diplotypeLoop['allele2']==starAllelesListTwoBasic[1]):
                                                # print '\njson file name:',dosingGuidelinesJsonFile
                                                # print strip_tags(annsLoop['groups'][0]['term']+levelOfEvidence+'  : '+annsLoop['textHtml'])
                                                # jsonSnp[annsLoop['groups'][0]['term']] = strip_tags(annsLoop['textHtml'])

                                                drug = parsedJsonFile['relatedDrugs'][0]['name']
                                                if annsLoop['groups'][0]['term']=='Recommendations':
                                                    rec = strip_tags(annsLoop['textHtml'])
                                                if annsLoop['groups'][0]['term']=='Recommendations':
                                                    levelOfEvidence = 'Level of Evidence: '+annsLoop['strength']['term']
                                                if (drug and rec and levelOfEvidence):
                                                    jsonSnp['pharmgkb']['drugrecommendations'].append({'haplotypes':diplotypeLoop['allele1']+diplotypeLoop['allele2'],'drug':drug,'recommendation':rec,'levelOfEvidence':levelOfEvidence})
                                    else:
                                        print 'no diplotypes at all in json file! but there are some guides in the json file!'
                            else:
                                pass
        else:
            print 'not enough basic star alleles means no PGx evidence for this variant.'
    for haplotype in haplottypeListComplete:
            jsonSnp['pharmgkb']['haplotypes'].append(haplotype)
    # print json.dumps(jsonSnp, indent=4, sort_keys=True)
    return jsonSnp

def get_mapping():
    mapping = {
        "pharmgkb": {
            "drugrecommendations": {
                "properties": {
                    "haplotypes": {
                        "type": "string",
                    },
                    "levelOfEvidence": {
                        "type": "string",
                    },
                    "drug": {
                        "type": "string"
                    },
                    "recommendation": {
                        "type": "string"
                    }
                }
            },
            "haplotypes": {
                "type": "string",
            }
        }
    }
    return mapping
